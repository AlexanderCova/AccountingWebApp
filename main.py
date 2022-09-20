from flask import Flask, render_template,request, redirect, url_for, send_file
import firebase_admin
from firebase_admin import credentials, firestore, storage
import pandas as pd
from werkzeug.utils import secure_filename
from flask_uploads import IMAGES, UploadSet, configure_uploads
import os
import pyshorteners
from openpyxl import Workbook
from datetime import date

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred, {'storageBucket' : 'obrienbbq-accounting.appspot.com'})

db = firestore.client()

app = Flask(__name__)


type_tiny = pyshorteners.Shortener()

temp_data = []


bucket = storage.bucket()
photos = UploadSet("photos", IMAGES)


app.config["UPLOADED_PHOTOS_DEST"] = "static/img"
app.config["SECRET_KEY"] = os.urandom(24)
configure_uploads(app, photos)

collections = ["equiptment", "food", "packaging"]

@app.route('/')
def hello_world():

	data = []

	headings = ("Name", "Price", "Category", "Reciept URL", "Truck")
	
	if request.args.get('category') != None and request.args.get('category') != "":
		collections = [request.args.get('category')]
	else:
		collections = ["equiptment", "food", "packaging"]

	if request.args.get('item-name') != None and request.args.get('item-name') != "":
		for i in collections:
			docs = db.collection(i).stream()

			for doc in docs:
				if request.args.get('item-name') in doc.to_dict()['name']:
					data.append((doc.to_dict()['name'], doc.to_dict()['price'], doc.to_dict()['category']))

	else:
		for i in collections:
			docs = db.collection(i).stream()

			for doc in docs:
				data.append((doc.to_dict()['name'], doc.to_dict()['price'], doc.to_dict()['category'], type_tiny.tinyurl.short(doc.to_dict()["url_to_reciept"]), doc.to_dict()["truck"]))



	return render_template("index.html", headings=headings, data=data)


@app.route("/add-item", methods=["GET", "POST"])
def add_item():

	if request.method == "POST":
		temp_data.append((request.form.get('name'), request.form.get('price'), request.form.get('category'), request.form.get("truck")))


	headings = ("Name", "Price", "Category", "Truck")

	return render_template("add-item.html", data=temp_data, headings=headings)


@app.route("/submit-items", methods=["POST"])
def submit_items():

	picture = request.files.get('img')
	photos.save(picture)

	blob = bucket.blob("static/img/"+picture.filename)
	blob.upload_from_filename("static/img/"+picture.filename)
	blob.make_public()

	for i in temp_data:
		db.collection(i[2]).add({
			'name': i[0],
			'price': i[1],
			'category': i[2],
			'url_to_reciept': blob.public_url,
			'truck': i[3]
			})

	os.remove("static/img/"+picture.filename)

	return redirect(url_for('hello_world'))


@app.route('/export')
def export():
	wb = Workbook()
	wb.create_sheet("Popup")
	wb["Sheet"].title = "Mac Shack"

	mac_shack_sheet = wb["Mac Shack"]
	popup_sheet = wb["Popup"]

	row_start = 3
	col_start = 2

	i = 0
	for item in collections:
		mac_shack_docs = db.collection(item).where("truck", "==", "Mac Shack").stream()
		

		

		for doc in mac_shack_docs:
			mac_shack_sheet.cell(row_start+i, col_start).value = doc.to_dict()['name']
			mac_shack_sheet.cell(row_start+i, col_start+1).value = doc.to_dict()['price']
			mac_shack_sheet.cell(row_start+i, col_start + 2).value = doc.to_dict()['category']
			mac_shack_sheet.cell(row_start+i, col_start+3).value = doc.to_dict()['url_to_reciept']
			i += 1


	i = 0

	for item in collections:
		popup_docs = db.collection(item).where("truck", "==", "Popup").stream()
		

		for doc in popup_docs:
			popup_sheet.cell(row_start+i, col_start).value = doc.to_dict()['name']
			popup_sheet.cell(row_start+i, col_start+1).value = doc.to_dict()['price']
			popup_sheet.cell(row_start+i, col_start + 2).value = doc.to_dict()['category']
			popup_sheet.cell(row_start+i, col_start+3).value = type_tiny.tinyurl.short(doc.to_dict()["url_to_reciept"])
			i += 1


	spreadsheet_path = str(date.today()) + "_export.xlsx"

	wb.save(spreadsheet_path)

	return send_file(spreadsheet_path)



